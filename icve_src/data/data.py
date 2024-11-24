import os

from openpyxl import load_workbook


class User:
    def __init__(self, account, password):
        self.account = account
        self.password = password

    def __repr__(self):
        return f"User(username='{self.account}', password='{self.password}')"

def read_excel():
    # 获取当前文件的绝对路径
    current_directory = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(current_directory, 'userData.xlsx')

    # 尝试加载工作簿
    try:
        wb = load_workbook(path)
        ws = wb.active
        # 使用生成器返回行数据
        yield from ws.iter_rows(min_row=2, values_only=True)

    except FileNotFoundError:
        print(f"错误：未找到文件 {path}")
    except Exception as e:
        print("发生错误：", e)


def get_user():
    list_user = []
    for user in read_excel():
        list_user.append(User(*user))
    print(list_user)
    return list_user

